from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import timedelta
import json
import csv
from io import StringIO
import logging

# Set up logging
logger = logging.getLogger(__name__)

from .models import StockPrediction, PredictionCache
from .serializers import (
    StockPredictionSerializer, 
    PredictionRequestSerializer,
    StockListSerializer
)
from .prediction_engine import StockPredictionEngine, AVAILABLE_STOCKS


@api_view(['GET'])
@permission_classes([AllowAny])
def get_available_stocks(request):
    """Get list of available stock symbols for prediction"""
    try:
        serializer = StockListSerializer({'stocks': AVAILABLE_STOCKS})
        return Response(serializer.data)
    except Exception as e:
        return Response(
            {'error': f'Failed to fetch stocks: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def generate_prediction(request):
    """Generate stock price prediction"""
    try:
        logger.info(f"Received prediction request: {request.data}")
        
        # Validate request data
        serializer = PredictionRequestSerializer(data=request.data)
        if not serializer.is_valid():
            logger.error(f"Validation errors: {serializer.errors}")
            return Response(
                {'error': 'Invalid request data', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        symbol = serializer.validated_data['symbol'].upper()
        timeframe = serializer.validated_data['timeframe']
        
        logger.info(f"Processing prediction for {symbol} with timeframe {timeframe}")
        
        # Check if symbol is available
        if symbol not in AVAILABLE_STOCKS:
            logger.warning(f"Symbol {symbol} not in available stocks list")
            return Response(
                {'error': f'Stock {symbol} is not available for prediction. Available stocks: {", ".join(AVAILABLE_STOCKS)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check cache first
        cache_key = f"{symbol}_{timeframe}"
        cached_prediction = PredictionCache.objects.filter(
            cache_key=cache_key,
            expires_at__gt=timezone.now()
        ).first()
        
        if cached_prediction:
            logger.info(f"Returning cached prediction for {symbol}")
            return Response(cached_prediction.cache_data)
        
        # Generate new prediction
        logger.info(f"Generating new prediction for {symbol}")
        prediction_engine = StockPredictionEngine()
        prediction_result = prediction_engine.predict(symbol, timeframe)
        
        logger.info(f"Prediction generated successfully for {symbol}")
        
        # Save prediction to database
        try:
            prediction = StockPrediction.objects.create(
                symbol=symbol,
                timeframe=timeframe,
                historical_data=prediction_result['historical_data'],
                predicted_data=prediction_result['predicted_data'],
                trend_direction=prediction_result['trend_direction'],
                confidence_score=prediction_result['confidence_score'],
                recommendation=prediction_result['recommendation'],
                model_used=prediction_result['model_used'],
                current_price=prediction_result['current_price'],
                predicted_end_price=prediction_result['predicted_end_price'],
                expected_return=prediction_result['expected_return']
            )
            logger.info(f"Prediction saved to database with ID: {prediction.id}")
        except Exception as db_error:
            logger.error(f"Database save error: {str(db_error)}")
            # Continue without saving to database if there's an error
        
        # Cache the prediction for 1 hour
        try:
            cache_expiry = timezone.now() + timedelta(hours=1)
            PredictionCache.objects.create(
                cache_key=cache_key,
                cache_data=prediction_result,
                expires_at=cache_expiry
            )
            logger.info(f"Prediction cached for {symbol}")
        except Exception as cache_error:
            logger.error(f"Cache error: {str(cache_error)}")
            # Continue without caching if there's an error
        
        # Add prediction ID to response if available
        if 'prediction' in locals():
            prediction_result['prediction_id'] = prediction.id
        
        logger.info(f"Returning prediction result for {symbol}")
        return Response(prediction_result)
        
    except ValueError as e:
        error_msg = f"Validation error: {str(e)}"
        logger.error(f"ValueError: {error_msg}")
        return Response(
            {'error': error_msg, 'type': 'validation_error'},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        error_msg = f'Prediction generation failed: {str(e)}'
        logger.error(f"Exception: {error_msg}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response(
            {'error': error_msg, 'type': 'server_error', 'details': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

from django.http import HttpResponse

# @api_view(['GET'])
# @permission_classes([AllowAny])
# def download_prediction_csv(request, symbol, timeframe):
#     """Download prediction data as CSV"""
#     try:
#         prediction = StockPrediction.objects.filter(
#             symbol=symbol.upper(),
#             timeframe=timeframe
#         ).order_by('-created_at').first()
        
#         if not prediction:
#             return HttpResponse(
#                 f'No prediction found for {symbol} with timeframe {timeframe}',
#                 status=status.HTTP_404_NOT_FOUND
#             )
        
#         # Use StringIO and csv writer
#         output = StringIO()
#         writer = csv.writer(output, lineterminator='\n')  # use \n for Excel
        
#         # Write header
#         writer.writerow(['Date', 'Historical Price', 'Predicted Price', 'Type'])
        
#         # Historical data
#         for date, price in zip(prediction.get_historical_dates(), prediction.get_historical_prices()):
#             writer.writerow([date, price, '', 'Historical'])
        
#         # Predicted data
#         for date, price in zip(prediction.get_predicted_dates(), prediction.get_predicted_prices()):
#             writer.writerow([date, '', price, 'Predicted'])
        
#         # Add a blank line before summary
#         writer.writerow([])
#         writer.writerow(['Summary'])
#         writer.writerow(['Symbol', symbol.upper()])
#         writer.writerow(['Timeframe', timeframe])
#         writer.writerow(['Current Price', prediction.current_price])
#         writer.writerow(['Predicted End Price', prediction.predicted_end_price])
#         writer.writerow(['Expected Return', f"{prediction.expected_return}%"])
#         writer.writerow(['Trend Direction', prediction.trend_direction])
#         writer.writerow(['Confidence Score', f"{prediction.confidence_score * 100:.1f}%"])
#         writer.writerow(['Recommendation', prediction.recommendation])
#         writer.writerow(['Model Used', prediction.model_used])
#         writer.writerow(['Generated At', prediction.created_at.strftime('%Y-%m-%d %H:%M:%S')])
        
#         # Return HttpResponse with proper encoding and BOM
#         response = HttpResponse(
#             output.getvalue(),
#             content_type='text/csv; charset=utf-8-sig'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{symbol}_{timeframe}_prediction.csv"'
        
#         return response

#     except Exception as e:
#         return HttpResponse(
#             f'CSV download failed: {str(e)}',
#             status=status.HTTP_500_INTERNAL_SERVER_ERROR
#         )

import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from prediction.models import StockPrediction
from rest_framework.decorators import api_view

@api_view(['GET'])
def download_prediction_xlsx(request, symbol, timeframe):
    prediction = StockPrediction.objects.filter(
        symbol=symbol.upper(),
        timeframe=timeframe
    ).order_by('-created_at').first()
    
    if not prediction:
        return HttpResponse("No prediction found", status=404)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{symbol}_{timeframe}"
    
    # Header
    headers = ['Date', 'Historical Price', 'Predicted Price', 'Type']
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
    
    # Historical data
    for date, price in zip(prediction.get_historical_dates(), prediction.get_historical_prices()):
        ws.append([date, price, '-', 'Historical'])
    
    # Predicted data
    for date, price in zip(prediction.get_predicted_dates(), prediction.get_predicted_prices()):
        ws.append([date, '-', price, 'Predicted'])
    
    # Blank line before summary
    ws.append([])
    
    # Summary section
    summary_title = ['Summary']
    ws.append(summary_title)
    # Make "Summary" bold and centered
    summary_cell = ws['A' + str(ws.max_row)]
    summary_cell.font = Font(bold=True)
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    summary_rows = [
        ['Symbol', symbol.upper()],
        ['Timeframe', timeframe],
        ['Current Price', prediction.current_price],
        ['Predicted End Price', prediction.predicted_end_price],
        ['Expected Return', f"{prediction.expected_return}%"],
        ['Trend Direction', prediction.trend_direction],
        ['Confidence Score', f"{prediction.confidence_score * 100:.1f}%"],
        ['Recommendation', prediction.recommendation],
        ['Model Used', prediction.model_used],
        ['Generated Date', prediction.created_at.strftime('%Y-%m-%d')],
        ['Generated Time', prediction.created_at.strftime('%H:%M:%S')]
    ]
    
    for row in summary_rows:
        ws.append(row)
    
    # Center align all cells and auto-fit columns
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # add some padding
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={symbol}_{timeframe}_prediction.xlsx'
    wb.save(response)
    
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def get_prediction_history(request, symbol):
    """Get prediction history for a specific stock"""
    try:
        predictions = StockPrediction.objects.filter(
            symbol=symbol.upper()
        ).order_by('-created_at')[:10]  # Last 10 predictions
        
        serializer = StockPredictionSerializer(predictions, many=True)
        return Response(serializer.data)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to fetch prediction history: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_stock_info(request, symbol):
    """Get basic stock information"""
    try:
        from .prediction_engine import StockPredictionEngine
        
        prediction_engine = StockPredictionEngine()
        stock_data = prediction_engine.get_stock_data(symbol, period='1mo')
        
        if stock_data.empty:
            return Response(
                {'error': f'No data found for {symbol}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        current_price = stock_data['Close'].iloc[-1]
        price_change = stock_data['Close'].iloc[-1] - stock_data['Close'].iloc[-2]
        price_change_pct = (price_change / stock_data['Close'].iloc[-2]) * 100
        
        stock_info = {
            'symbol': symbol.upper(),
            'current_price': round(current_price, 2),
            'price_change': round(price_change, 2),
            'price_change_pct': round(price_change_pct, 2),
            'volume': int(stock_data['Volume'].iloc[-1]),
            'high_52w': round(stock_data['High'].max(), 2),
            'low_52w': round(stock_data['Low'].min(), 2),
            'last_updated': stock_data.index[-1].strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return Response(stock_info)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to fetch stock info: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([AllowAny])
def clear_prediction_cache(request):
    """Clear expired prediction cache (admin function)"""
    try:
        expired_cache = PredictionCache.objects.filter(
            expires_at__lt=timezone.now()
        )
        count = expired_cache.count()
        expired_cache.delete()
        
        return Response({
            'message': f'Cleared {count} expired cache entries',
            'cleared_count': count
        })
        
    except Exception as e:
        return Response(
            {'error': f'Failed to clear cache: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
